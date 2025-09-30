import base64
import io
import json
import os
import shutil
import tempfile
import time
import zipfile
import zlib
from dataclasses import asdict, dataclass, field
from glob import glob
from typing import Any, Dict, List, Optional, Tuple

import cv2
import nibabel as nib
import numpy as np
import pandas as pd
import pydicom
import SimpleITK as sitk
import torch
import torch.nn as nn
import torch.nn.functional as F
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from scipy import ndimage as ndi
from scipy.ndimage import gaussian_filter, label
from scipy.stats import norm
from sklearn.metrics import roc_auc_score
import matplotlib.pyplot as plt


from dinov3.models.vision_transformer import vit_base



MODELS = [
        ("covid19",      r"./weights/model"),
        ("aneurysm_TAA", r"./weights/model"),
        ("pleural_eff",  r"./weights/model"),
    ]

RECOMMENDED_PROFILES = {
    "fast_debug": {
        "inp.k_25d": 3,
        "inp.target_hw": (224,224),
        "train.mil_samples_per_patient": 32,
    },
    "balanced": {
        "inp.k_25d": 5,
        "inp.target_hw": (384,384),
        "train.mil_samples_per_patient": 128,
    },
    "high_quality": {
        "inp.k_25d": 7,
        "inp.target_hw": (512,512),
        "train.mil_samples_per_patient": 256,
    }
}

_POSITIVE_TOKENS = [
    "CHEST","THORAX","LUNG","THORACIC","CHEST CT","THORAX CT","CT CHEST","THORAKS","CHEST/ABD",
    "КГК","ГРУД","ЛЁГК","ЛЕГК","ТОРАК","ГРУДНАЯ КЛЕТКА","ОРГАНЫ ГРУДНОЙ КЛЕТКИ",
]
_NEGATIVE_TOKENS = [
    "HEAD","BRAIN","NEURO","NECK","C-SPINE","L-SPINE","THORACIC SPINE","ABDOMEN","PELVIS","KNEE","CARDIAC CT",
    "ГОЛОВ","МОЗГ","ШЕЯ","ПОЗВОН","ЖИВОТ","ТАЗ","КОЛЕН",
]


class PreprocessConfig:
    def __init__(self):
        self.iso_spacing: Tuple[float, float, float] = (1.0, 1.0, 1.0)  # изотропный ресемплинг в мм
        self.trim_min_frac_area: float = 0.05
        self.trim_run_len: int = 6
        self.trim_margin: int = 8
        self.windows: List[Tuple[float, float]] = ((-600, 1500), (40, 400))  # (center, width): лёгочное и медиастинальное
        self.use_bone_window: bool = False  # добавить костное окно (+300,1500)

class InputConfig:
    def __init__(self):
        self.use_25d: bool = True
        self.k_25d: int = 5               # нечётное (3/5/7)
        self.target_hw: Tuple[int, int] = (384, 384)

class ModelConfig:
    def __init__(self):
        self.backbone_type: str = "vit_2d"   # "vit_2d" | "resnet3d_medicalnet" | "cnn3d_custom"
        self.in_chans: int = 10              # окна×k (например 2 окна × 5 k = 10)
        self.feat_dim: int = 768             # выходной размер признаков ViT/encoder
        self.num_classes: int = 1

class EvalConfig:
    def __init__(self):
        self.device: str = "cuda" if torch.cuda.is_available() else "cpu"
        self.mil_samples_per_patient: int = 128  # M: сколько центральных срезов брать у пациента на шаг

class FullConfig:
    def __init__(self):
        self.prep: PreprocessConfig = PreprocessConfig()
        self.inp: InputConfig = InputConfig()
        self.model: ModelConfig = ModelConfig()
        self.eval: EvalConfig = EvalConfig()

    def as_dict(self):
        return {k: v for k, v in self.__class__.__dict__.items()
                if not k.startswith("__") and not callable(v)}

class CTPreprocessor:
    """Чтение DICOM → HU → ресемплинг → mask → trim → окна → отдаёт объём и служебные маски."""
    def __init__(self, cfg: PreprocessConfig):
        self.cfg = cfg


    def _is_scout(self, ds) -> bool:
        """Эвристика: локалайзер/скаут — 1-2 кадра, ImageType содержит LOCALIZER/SCOUT, или SeriesDescription намекает."""
        try:
            it = [s.upper() for s in getattr(ds, "ImageType", [])] if hasattr(ds, "ImageType") else []
            desc = str(getattr(ds, "SeriesDescription", "")).upper()
            proto = str(getattr(ds, "ProtocolName", "")).upper()
            tokens = ("LOCALIZER", "LOCALISER", "SCOUT", "TOPOGRAM", "TOPOGRAM")
            if any(t in it for t in tokens): return True
            if any(t in desc for t in tokens): return True
            if any(t in proto for t in tokens): return True
        except Exception:
            pass
        return False

    def _series_len_estimate(self, files_with_hdrs: list) -> int:
        """Грубая оценка числа срезов: считаем уникальные позиции Z.
        Для multi-frame (1 файл) вернётся 1 — позже проверим фактический Z через чтение."""
        zs = []
        for _, ds in files_with_hdrs:
            ipp = getattr(ds, "ImagePositionPatient", None)
            if ipp and len(ipp) == 3:
                zs.append(float(ipp[2]))
            else:
                inst = getattr(ds, "InstanceNumber", None)
                zs.append(float(inst) if inst is not None else 0.0)
        return len(set(zs)) if zs else len(files_with_hdrs)

    def _rank_series(self, series_dict: dict) -> list:
        """Возвращает список (uid, files_with_hdrs) в порядке убывания качества, с фильтром на CT/монохром и без scout."""
        candidates = []
        for uid, files in series_dict.items():
            # фильтруем «битые» хедеры
            files = [(p, ds) for (p, ds) in files if ds is not None]
            if not files:
                continue
            # CT + монохром
            files_ct = [(p, ds) for (p, ds) in files if is_ct_monochrome(ds)]
            if not files_ct:
                continue
            # отсеиваем очевидные локалайзеры
            if all(self._is_scout(ds) for (_, ds) in files_ct):
                continue

            n_est = self._series_len_estimate(files_ct)
            score = series_score(files_ct) + 0.1 * n_est  # чуть поднимаем длинные серии
            candidates.append((uid, files_ct, score))

        # если после отсева пусто — вернём все CT-моно серии без фильтра scout
        if not candidates:
            for uid, files in series_dict.items():
                files_ct = [(p, ds) for (p, ds) in files if is_ct_monochrome(ds)]
                if files_ct:
                    n_est = self._series_len_estimate(files_ct)
                    score = series_score(files_ct) + 0.1 * n_est
                    candidates.append((uid, files_ct, score))

        # сортировка по score ↓
        candidates.sort(key=lambda t: t[2], reverse=True)
        return [(uid, files) for (uid, files, _) in candidates]

    def load_dicom_series(self, dicom_dir: str, min_slices: int = 16, skip_if_bad: bool = False) -> sitk.Image:
        """
        Выбирает лучшую серию:
        - игнорирует scout/localizer,
        - пробует серии по рангу до тех пор, пока не найдёт достаточную по Z,
        - поддерживает Enhanced CT (один файл с множеством кадров).
        Если ничего подходящего — либо кидает RuntimeError, либо (при skip_if_bad=True) поднимает ValueError для обработки выше.
        """
        series_dict = collect_series(dicom_dir)
        if not series_dict:
            print(f"Нет DICOM-файлов в: {dicom_dir}")
            return None

        ranked = self._rank_series(series_dict)
        if not ranked:
            if skip_if_bad:
                print(f"В {dicom_dir} нет пригодных CT/моно серий")
                return None
            print(f"Не найдено CT/моно серий в: {dicom_dir}")
            return None

        # Перебираем кандидатов до успеха
        last_err = None
        for uid, files_with_hdrs in ranked:
            file_list = self._sort_slices(files_with_hdrs)  # список путей в серии (может быть из 1 файла)
            try:
                # Быстрая проверка: если файлов мало, это может быть Enhanced CT — читаем и смотрим реальный Z
                reader = sitk.ImageSeriesReader()
                reader.SetFileNames(file_list)
                img = reader.Execute()  # SimpleITK сам применит slope/intercept
                size = img.GetSize()    # (X, Y, Z)
                Z = int(size[2])
                if Z < min_slices:
                    # слишком мало срезов — похоже на локалайзер или неполную серию
                    last_err = RuntimeError(f"Серия {uid}: Z={Z} < min_slices={min_slices}")
                    continue
                return img  # успех
            except Exception as e:
                last_err = e
                continue

        return None

    def resample_iso(self, img: sitk.Image) -> sitk.Image:
        in_spacing = img.GetSpacing()
        in_size = img.GetSize()
        out_spacing = self.cfg.iso_spacing
        out_size = [int(round(in_size[i] * (in_spacing[i] / out_spacing[i]))) for i in range(3)]
        res = sitk.ResampleImageFilter()
        res.SetInterpolator(sitk.sitkLinear)
        res.SetOutputSpacing(out_spacing)
        res.SetSize(out_size)
        res.SetOutputDirection(img.GetDirection())
        res.SetOutputOrigin(img.GetOrigin())
        return res.Execute(img)

    @staticmethod
    def sitk_to_numpy(img: sitk.Image):
        # Узнаём число компонент на пиксель (сколько "каналов")
        ncomp = img.GetNumberOfComponentsPerPixel()  # 1 для обычного КТ, 3 для RGB
        arr = sitk.GetArrayFromImage(img)  # может быть [Z,Y,X] или [Z,Y,X,3]
        arr = np.asarray(arr)

        # Приводим к [Z,Y,X]
        if arr.ndim == 4 and arr.shape[-1] == 1:
            arr = arr[..., 0]
        elif arr.ndim == 4 and arr.shape[-1] == 3:
            # Проверим, одинаковые ли каналы (часто 3 одинаковых)
            if np.allclose(arr[...,0], arr[...,1]) and np.allclose(arr[...,1], arr[...,2]):
                arr = arr[...,0]
            else:
                # Перевод в серый: luminance
                arr = (0.2989*arr[...,0] + 0.5870*arr[...,1] + 0.1140*arr[...,2]).astype(arr.dtype)


        spacing = img.GetSpacing()
        origin  = img.GetOrigin()
        direction = img.GetDirection()
        return arr.astype(np.float32), spacing, origin, direction

    @staticmethod
    def body_mask(vol: np.ndarray) -> np.ndarray:
        
        vol = np.asarray(vol)
        if vol.ndim == 4 and vol.shape[-1] == 1:
            vol = vol[..., 0]
        assert vol.ndim == 3, f"body_mask: ожидался 3D, got {vol.shape}"
        mask = vol > -500
        mask = ndi.binary_closing(mask, iterations=2)
        mask = ndi.binary_fill_holes(mask)
        lab, n = ndi.label(mask)
        if n > 0:
            sizes = ndi.sum(mask, lab, index=np.arange(1, n+1))
            k = 1 + int(np.argmax(sizes))
            mask = (lab == k)
        return mask

    @staticmethod
    def lung_mask_simple(vol: np.ndarray, body_mask: np.ndarray) -> np.ndarray:
        lungs = (vol < -350) & body_mask
        lungs = ndi.binary_opening(lungs, iterations=2)
        lungs = ndi.binary_closing(lungs, iterations=1)
        return lungs

    def auto_trim_z(self, lung_mask: np.ndarray) -> Tuple[int, int]:
        lung_mask = np.asarray(lung_mask)
        if lung_mask.ndim == 4 and lung_mask.shape[-1] == 1:
            lung_mask = lung_mask[..., 0]
        # if lung_mask.ndim != 3:
        #     raise ValueError(f"auto_trim_z: ожидался 3D, got {lung_mask.shape}")
        Z, Y, X = lung_mask.shape
        areas = lung_mask.reshape(Z, -1).sum(axis=1).astype(np.float32)
        max_area = float(areas.max()) if areas.max() > 0 else 1.0
        thr = max_area * self.cfg.trim_min_frac_area
        valid = areas >= thr
        best_s, best_e, cur_s = None, None, None
        for i in range(Z):
            if valid[i] and cur_s is None:
                cur_s = i
            end = (not valid[i]) or (i == Z - 1)
            if end and cur_s is not None:
                e = i if not valid[i] else i
                if (e - cur_s + 1) >= self.cfg.trim_run_len:
                    if best_s is None or (e - cur_s) > (best_e - best_s):
                        best_s, best_e = cur_s, e
                cur_s = None
        if best_s is None:
            return 0, Z
        s = max(0, best_s - self.cfg.trim_margin)
        e = min(Z, best_e + self.cfg.trim_margin + 1)
        return s, e

    @staticmethod
    def window_apply(vol: np.ndarray, center: float, width: float) -> np.ndarray:
        low, high = center - width/2.0, center + width/2.0
        v = np.clip(vol, low, high)
        v = (v - low) / max(1e-6, (high - low))
        return v.astype(np.float32)

    def make_windows(self, vol: np.ndarray) -> np.ndarray:
        # Возвращает [Z, C, Y, X]
        channels = []
        for (c, w) in self.cfg.windows:
            channels.append(self.window_apply(vol, c, w))
        if self.cfg.use_bone_window:
            channels.append(self.window_apply(vol, 300.0, 1500.0))
        ch = np.stack(channels, axis=1)  # [Z, C, Y, X]
        return ch

    @staticmethod
    def _sort_slices(files_with_hdrs: list) -> list:
        """
        Стабильная сортировка по координате Z (ImagePositionPatient[2]) или InstanceNumber.
        Возвращает список путей.
        """
        def z_of(ds):
            ipp = getattr(ds, "ImagePositionPatient", None)
            if ipp and len(ipp) == 3:
                return float(ipp[2])
            inst = getattr(ds, "InstanceNumber", None)
            return float(inst) if inst is not None else 0.0

        files_with_hdrs = [(p, ds) for (p, ds) in files_with_hdrs if ds is not None]
        files_with_hdrs.sort(key=lambda t: z_of(t[1]))
        return [p for (p, _) in files_with_hdrs]


    def __call__(self, dicom_dir: str, min_slices: int = 16):
        """
        Мягкая загрузка и препроцесс серии:
        - выбираем подходящую серию (внутри load_dicom_series обрабатываются scout/multiframe),
        - при любых проблемах/аномалиях возвращаем None (а не исключение),
        - ресемплируем до iso_spacing, строим маски, auto-trim по Z,
        - применяем окна и возвращаем [Z,C,Y,X] + служебные маски/метаданные.

        Возвращает:
        Tuple[np.ndarray, np.ndarray, np.ndarray, dict]  -> (vol_ch, lungs, body, meta)
        либо None, если серию пропускаем.
        """
        try:
            img = self.load_dicom_series(dicom_dir, min_slices=min_slices, skip_if_bad=True)
            if img is None:
                print(f"[WARN] {dicom_dir}: не удалось выбрать серию (img=None). Пропуск.")
                return None
        except Exception as e:
            print(f"[WARN] {dicom_dir}: ошибка выбора серии: {e}. Пропуск.")
            return None

        try:
            # 2) Ресемпл → изотропный шаг
            img = self.resample_iso(img)

            # 3) В HU numpy и базовые проверки
            vol, spacing, origin, direction = self.sitk_to_numpy(img)  # vol: [Z,Y,X] float32 (HU)
            if not np.isfinite(vol).all():
                vol = np.nan_to_num(vol, nan=-1000.0, posinf=2000.0, neginf=-1024.0)

            Z0 = int(vol.shape[0])
            if Z0 < 3:
                print(f"[WARN] {dicom_dir}: слишком мало срезов после чтения (Z={Z0}). Пропуск.")
                return None

            # 4) Маски тела/лёгких
            body = self.body_mask(vol)                     # [Z,Y,X] bool
            lungs = self.lung_mask_simple(vol, body)       # [Z,Y,X] bool

            # Если лёгкие не нашлись (например, КТ не грудной клетки) — попробуем тримминг по body
            trim_basis = lungs if lungs.sum() > 0 else body

            # 5) Auto-trim по Z
            z0, z1 = self.auto_trim_z(trim_basis)          # полуинтервал [z0, z1)
            if (z1 - z0) < 3:
                print(f"[WARN] {dicom_dir}: после trim осталось {z1 - z0} срезов. Пропуск.")
                return None

            vol   = vol[z0:z1]
            body  = body[z0:z1]
            lungs = lungs[z0:z1]

            # 6) Применяем окна → [Z, Cwin, Y, X]
            vol_ch = self.make_windows(vol)

            # 7) Метаданные (полезно логировать причины и состояния)
            meta = dict(
                spacing=tuple(map(float, spacing)),
                origin=tuple(map(float, origin)),
                direction=tuple(map(float, direction)),
                z_crop=(int(z0), int(z1)),
                z_before=int(Z0),
                z_after=int(vol.shape[0]),
                lung_voxels=int(lungs.sum()),
                body_voxels=int(body.sum()),
                dicom_dir=str(dicom_dir),
            )

            return vol_ch, lungs, body, meta

        except Exception as e:
            print(f"[WARN] {dicom_dir}: ошибка препроцесса: {e}. Пропуск.")
            return None

class ChestCTGate:
    """
    Возвращает dict:
      {
        'is_chest': bool,
        'score': float (0..1),
        'reason': str,
        'features': {...}  # полезные измерения
      }
    Порог по score регулируется.
    """
    def __init__(self,
                 meta_weight: float = 0.35,
                 hu_weight: float = 0.65,
                 score_threshold: float = 0.55,
                 min_slices: int = 80,
                 z_extent_mm_range=(180, 450),
                 spacing_xy_max_mm=1.5,
                 lung_air_hu=(-950, -500),
                 soft_tissue_hu=(-100, 300),
                 bone_hu=(300, 2000)):
        self.meta_w = meta_weight
        self.hu_w = hu_weight
        self.thr = score_threshold
        self.min_slices = min_slices
        self.z_extent_range = z_extent_mm_range
        self.spacing_xy_max = spacing_xy_max_mm
        self.lung_air_hu = lung_air_hu
        self.soft_hu = soft_tissue_hu
        self.bone_hu = bone_hu

    # ---------- метаданные ----------
    def _meta_score(self, headers):
        """
        Простая лексическая проверка BodyPartExamined / SeriesDescription / ProtocolName.
        +1 за положительный токен, -1 за отрицательный, суммируем и squash → [0..1]
        """
        votes = 0
        for tag in ("BodyPartExamined", "SeriesDescription", "ProtocolName"):
            v = getattr(headers, tag, None)
            if v is None: 
                continue
            s = str(v).upper()
            if any(t in s for t in _POSITIVE_TOKENS): votes += 1
            if any(t in s for t in _NEGATIVE_TOKENS): votes -= 1
        # squash в [0..1]
        return 1 / (1 + np.exp(-votes))

    def _hu_geo_score(self, vol, sp):
        Z, Y, X = vol.shape
        sx, sy, sz = sp  # sitk: (x,y,z)

        # Базовые проверки геометрии
        geo_ok = 1.0
        # достаточное число срезов
        if Z < self.min_slices:
            geo_ok *= 0.6
        # разумный охват по Z
        z_extent_mm = Z * sz
        if not (self.z_extent_range[0] <= z_extent_mm <= self.z_extent_range[1]):
            geo_ok *= 0.6
        # XY spacing (чтобы не было 3-5 мм, характерных для scout/локалайзера)
        if max(sx, sy) > self.spacing_xy_max:
            geo_ok *= 0.6

        body = _body_mask(vol)
        lungs = _lung_mask(vol, body)

        total = float(body.sum()) + 1e-6

        # доля воздуха в лёгком диапазоне внутри тела
        air_lo, air_hi = self.lung_air_hu
        lung_like = ((vol >= air_lo) & (vol <= air_hi) & body).sum() / total

        # доля мягких тканей
        st_lo, st_hi = self.soft_hu
        soft_like = ((vol >= st_lo) & (vol <= st_hi) & body).sum() / total

        # доля костей
        b_lo, b_hi = self.bone_hu
        bone_like = ((vol >= b_lo) & (vol <= b_hi) & body).sum() / total

        # Проверка «двух лёгких» на средних срезах
        if Z >= 9:
            mid = slice(Z//3, 2*Z//3)
        else:
            mid = slice(0, Z)
        two_lungs_scores = []
        for z in range(*mid.indices(Z)):
            two_lungs_scores.append(_two_lungs_fraction(lungs[z]))
        two_lungs = float(np.median(two_lungs_scores) if two_lungs_scores else 0.0)

        # Эвристический скор HU/геометрии.
        # Лёгкое КТ обычно имеет:
        # - lung_like ~ 0.2-0.6
        # - soft_like ~ 0.3-0.6
        # - bone_like ~ 0.02-0.15
        # - two_lungs ~ 0.6-0.95
        # - geo_ok близко к 1
        # Сконструируем «идеал» и возьмём смешанную близость.
        def clamp01(x): return max(0.0, min(1.0, float(x)))

        # подочки (каждый в [0..1])
        s_lung = clamp01(1.0 - abs(lung_like - 0.35) / 0.35)     # best ~0.35
        s_soft = clamp01(1.0 - abs(soft_like - 0.45) / 0.45)     # best ~0.45
        s_bone = clamp01(1.0 - abs(bone_like - 0.07) / 0.07)     # best ~0.07
        s_two  = clamp01(two_lungs)                               # уже [0..1]
        s_geo  = clamp01(geo_ok)

        # сводный HU-скор (усреднение с весами)
        hu_score = (0.35*s_lung + 0.30*s_soft + 0.10*s_bone + 0.15*s_two + 0.10*s_geo)
        features = dict(
            Z=Z, spacing_xyz=(sx,sy,sz), z_extent_mm=z_extent_mm,
            lung_like=lung_like, soft_like=soft_like, bone_like=bone_like,
            two_lungs=two_lungs, geo_ok=geo_ok,
            s_lung=s_lung, s_soft=s_soft, s_bone=s_bone, s_two=s_two, s_geo=s_geo,
            hu_score=hu_score
        )
        return hu_score, features

    def is_chest_ct(self, dicom_dir: str):
        # 1) метаданные
        series = _read_series_headers(dicom_dir)
        if not series:
            return dict(is_chest=False, score=0.0, reason="no_dicom_series", features={})
        # берём первую попавшуюся шапку для текстовых тегов
        first_hdr = next(iter(series.values()))[0][1]
        meta_s = self._meta_score(first_hdr)

        # 2) объём (HU)
        try:
            paths = _best_series_paths(series)
            img = _read_image(paths)
            vol, sp, sz = _sitk_to_np(img)      # [Z,Y,X], HU
            hu_s, feats = self._hu_geo_score(vol, sp)
        except Exception as e:
            # если не смогли прочитать — решаем по метаданным
            final = self.meta_w*meta_s
            return dict(is_chest=(final >= self.thr), score=final,
                        reason=f"meta_only:{e}", features=dict(meta_score=meta_s))

        # 3) финальный скор
        final = float(self.meta_w*meta_s + self.hu_w*hu_s)

        # 4) вердикт
        is_chest = final >= self.thr
        reason = "ok" if is_chest else "below_threshold"
        feats.update(dict(meta_score=meta_s, final_score=final))
        return dict(is_chest=is_chest, score=final, reason=reason, features=feats)

class MedDINOv3Backbone(nn.Module):
    def __init__(self, in_chans: int, img_size: int = 384, ckpt_path: str = None):
        super().__init__()
        # Инициализация их ViT (параметры как в README)
        self.vit = vit_base(
            drop_path_rate=0.0, layerscale_init=1.0e-5, n_storage_tokens=4,
            qkv_bias=False, mask_k_bias=True
        )
        self.img_size = img_size
        self.patch = 16  # ViT-B/16
        self.embed_dim = 768

        # 2) Если хотим много каналов (2.5D: окна × k), адаптируем патч-эмбеддинг
        pe = self.vit.patch_embed.proj
        if pe.in_channels != in_chans:
            # Переопределим слой на нужное число входных каналов
            new_proj = nn.Conv2d(in_chans, pe.out_channels, kernel_size=pe.kernel_size,
                                stride=pe.stride, padding=pe.padding, bias=True)
            self.vit.patch_embed.proj = new_proj
        

        if ckpt_path:
            sd = torch.load(ckpt_path, map_location="cpu")
            if "teacher" in sd:
                sd = sd["teacher"]
            sd = {k.replace("backbone.", ""): v for k, v in sd.items()
                  if "ibot" not in k and "dino_head" not in k}

        # 2) Адаптируем веса из чекпойнта под новое число каналов
        w = sd.get("patch_embed.proj.weight", None)   # ожидаем [768, 3, 16, 16]
        b = sd.get("patch_embed.proj.bias", None)
        if w is not None:
            # усредняем по входному каналу и повторяем на C
            w_mean = w.mean(dim=1, keepdim=True)      # [768,1,16,16]
            w_rep  = w_mean.repeat(1, in_chans, 1, 1) # [768,C,16,16]
            sd["patch_embed.proj.weight"] = w_rep
            if b is None:
                sd["patch_embed.proj.bias"] = torch.zeros(w.shape[0])

        missing, unexpected = self.vit.load_state_dict(sd, strict=False)

        print("Loaded MedDINOv3, missing:", len(missing), "unexpected:", len(unexpected))

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        """
        Вход:  x  — [B, C, H, W] (C = окна×k)
        Выход: fmap — [B, D, h, w] (карта патч-признаков)
        """
        B, C, H, W = x.shape

        if (H, W) != (self.img_size, self.img_size):
            x = torch.nn.functional.interpolate(
                x, size=(self.img_size, self.img_size),
                mode="bilinear", align_corners=False
            )
        out = None
        if hasattr(self.vit, "forward_features"):
            out = self.vit.forward_features(x)
        else:
            out = self.vit(x)

        # 3) нормализуем к виду "последовательность токенов" или fmap
        if isinstance(out, dict):
            if "last_hidden_state" in out:
                tokens = out["last_hidden_state"]
            elif "x" in out:
                tokens = out["x"]
            elif "x_norm_patchtokens" in out:
                tokens = out["x_norm_patchtokens"]
                # можно сохранить cls отдельно, если нужно:

        # 4) если уже карта признаков — вернуть как есть
        if tokens.dim() == 4 and tokens.size(1) != 1:
            # [B, D, h, w]
            return tokens

        N = tokens.size(1)
        expected_with_cls = 1 + (self.img_size // self.patch) * (self.img_size // self.patch)
        if N == expected_with_cls:
            tokens = tokens[:, 1:, :]  # [B, N, D] без CLS
            N = tokens.size(1)

        # 6) склеиваем токены в карту [B, D, h, w]
        D = tokens.size(2)
        # стараемся угадать квадратную решётку патчей
        hw = int(round(N ** 0.5))
        if hw * hw != N:
            # если N не квадрат — попробуем вычислить из img_size/patch
            hw = self.img_size // self.patch

        fmap = tokens.transpose(1, 2).contiguous().view(B, D, hw, hw)  # [B, D, h, w]
        return fmap

class AttentionMILHead(nn.Module):
    """
    Attention MIL: H[n,D] -> w[n] -> z[D] -> logits[C]
    """
    def __init__(self, in_dim: int, num_classes: int = 1, dropout: float = 0.2):
        super().__init__()
        self.attn = nn.Sequential(nn.Linear(in_dim, 128), nn.Tanh(), nn.Linear(128, 1))
        self.cls = nn.Sequential(nn.Dropout(dropout), nn.Linear(in_dim, num_classes))

    def forward(self, H: torch.Tensor):
        # H: [n, D] (мешок срезов одного пациента)
        a = self.attn(H).squeeze(-1)         # [n]
        w = torch.softmax(a, dim=0)          # [n]
        z = (w.unsqueeze(-1) * H).sum(0)     # [D]
        logits = self.cls(z).unsqueeze(0)    # [1,C]
        return logits, w

class Builder2p5D:
    """Строит 2.5D стек: соседние срезы по Z упакованы в каналы + ресайз до target_hw."""
    def __init__(self, k: int = 5, target_hw: Tuple[int, int] = (384, 384)):
        self.k = k
        self.target_hw = target_hw

    def build_for_volume(self, vol_ch: np.ndarray) -> torch.Tensor:
        """
        vol_ch: [Z, Cw, H, W], где Cw = число окон
        Возвращает: тензор [Z, Cw*k, Ht, Wt] — один элемент на каждый центральный срез
        """
        Z, Cw, H, W = vol_ch.shape
        pad = self.k // 2
        vol_pad = np.pad(vol_ch, ((pad,pad),(0,0),(0,0),(0,0)), mode='edge')
        out = []
        for z in range(Z):
            block = vol_pad[z:z+self.k]              # [k, Cw, H, W]
            block = np.transpose(block, (1,0,2,3))   # [Cw, k, H, W]
            block = block.reshape(Cw*self.k, H, W)   # [Cw*k, H, W]
            t = torch.from_numpy(block)[None]        # [1, C, H, W]
            t = F.interpolate(t, size=self.target_hw, mode='bilinear', align_corners=False)
            out.append(t)
        out = torch.cat(out, dim=0)  # [Z, C, Ht, Wt]
        return out.float()

class TrainerMIL:
    def __init__(self, cfg: FullConfig, ckpt_path: str):
        self.cfg = cfg
        self.ckpt_path = ckpt_path

        self.device = cfg.eval.device

        if cfg.model.backbone_type == "vit_2d":
            self.backbone = MedDINOv3Backbone(
                in_chans=cfg.model.in_chans,
                img_size=cfg.inp.target_hw[0],
                ckpt_path=self.ckpt_path 
            ).to(self.device)
            self.is_2d = True

        self.head = AttentionMILHead(cfg.model.feat_dim, num_classes=cfg.model.num_classes, dropout=0.3)

        self.chest_gate = ChestCTGate(score_threshold=0.55)

        
        self.backbone.to(self.device)
        self.head.to(self.device)

        self.prep = CTPreprocessor(cfg.prep)
        self.b25d = Builder2p5D(cfg.inp.k_25d, cfg.inp.target_hw)

    
    def _overlay_cam(self, slice2d: np.ndarray, cam_small: np.ndarray, out_path: str):
        """
        slice2d: [H,W] в 0..1 (лучше окно lung/medi)
        cam_small: [h',w'] из feature-карт бэкбона
        """
        
        H, W = slice2d.shape
        cam = cv2.resize(cam_small, (W, H), interpolation=cv2.INTER_CUBIC)
        cam = (cam - cam.min()) / (cam.max() - cam.min() + 1e-6)
        heat = cv2.applyColorMap((cam*255).astype(np.uint8), cv2.COLORMAP_JET)[:, :, ::-1] / 255.0
        base = np.stack([slice2d]*3, axis=-1)
        out = 0.6*base + 0.4*heat
        out = np.clip(out, 0, 1)
        plt.imsave(out_path, out)


    @torch.no_grad()
    def _forward_patient_with_maps(self, dicom_dir: str):
        """
        Возвращает:
        logits: [1, C]
        w: [Z]               — attention веса по срезам
        fmaps: [Z, D, h, w]  — feature-карты бэкбона для каждого среза
        slices_for_vis: dict[z] -> 2D slice [H,W] в 0..1 (для наложения CAM)
        meta: dict
        """
        pack = self.prep(dicom_dir)
        if pack is None:
            return None
        vol_ch, lungs, body, meta = pack  # [Z, Cwin, H, W]

        Z, Cw, H, W = vol_ch.shape
        fmaps_chunks, feats_chunks = [], []
        z_order = []

        # фоновый срез для оверлея (берём 0-й канал первого окна)
        def _slice_for_vis(z):
            img = vol_ch[z, 0]  # [H, W], 0..1 (если нормализация в препроцессе)
            vmin, vmax = float(img.min()), float(img.max())
            if vmax > vmin:
                img = (img - vmin) / (vmax - vmin)
            else:
                img = np.zeros_like(img)
            return img

        # собираем батчами (порядок z может быть неотсортированным — фиксируем z_order)
        for x_batch, z_batch in self.make_slice_batches(vol_ch):  # x_batch: [B,C,H,W], z_batch: list[int]
            x_batch = x_batch.to(self.device, non_blocking=True)
            fmap = self.backbone(x_batch)      # [B, D, h, w]
            emb  = fmap.mean(dim=[2,3])        # [B, D]
            fmaps_chunks.append(fmap.cpu())
            feats_chunks.append(emb.cpu())
            z_order.extend(z_batch)

        # объединяем
        fmaps = torch.cat(fmaps_chunks, dim=0)   # сейчас в порядке z_order
        H_emb = torch.cat(feats_chunks, dim=0)   # [Z, D] в порядке z_order

        # восстанавливаем строгий Z-порядок
        order = np.argsort(np.array(z_order))
        fmaps = fmaps[order]                     # [Z, D, h, w] по z=0..Z-1
        H_emb = H_emb[order]                     # [Z, D]

        # MIL-голова
        logits, w = self.head(H_emb.to(self.device))   # logits [1,C], w [Z]

        # фон для CAM
        slices_for_vis = {z: _slice_for_vis(z) for z in range(Z)}

        return logits, w, fmaps, slices_for_vis, meta
 
    def make_slice_batches(self, vol_ch: np.ndarray, batch_slices: int = 16):
        """
        vol_ch: [Z,Cwin,H,W], строим 2.5D, затем отдаём батчи по срезам:
        yield x_batch:[B,C,H,W], z_batch:[int]
        """
        X = self.b25d.build_for_volume(vol_ch)  # [Z, Cw*k, Ht, Wt]
        Z = X.shape[0]
        for s in range(0, Z, batch_slices):
            e = min(Z, s + batch_slices)
            z_idx = list(range(s, e))
            yield X[s:e].float(), z_idx

class ModelAdapter:
    def __init__(self, device: str = "cuda", ckpt_path: str = None,
                 profile: str = "balanced"):
        self.device = device


        self.cfg = make_default_cfg()
        prof = RECOMMENDED_PROFILES.get(profile, {})

        self.chest_gate = ChestCTGate(score_threshold=0.55)

        if "inp.k_25d" in prof:
            self.cfg.inp.k_25d = prof["inp.k_25d"]

            self.cfg.model.in_chans = len(self.cfg.prep.windows) * self.cfg.inp.k_25d
        if "inp.target_hw" in prof:
            self.cfg.inp.target_hw = tuple(prof["inp.target_hw"])
        if "train.mil_samples_per_patient" in prof:
            self.cfg.eval.mil_samples_per_patient = prof["train.mil_samples_per_patient"]

        def print_cfg(cfg, indent=0):
            for k, v in cfg.__dict__.items():
                if hasattr(v, "__dict__"):
                    print("  " * indent + f"{k}:")
                    print_cfg(v, indent+1)
                else:
                    print("  " * indent + f"{k} = {v}")

        # print_cfg(self.cfg)

        self.trainer = TrainerMIL(self.cfg, ckpt_path=ckpt_path)
        if ckpt_path:
            sd = torch.load(ckpt_path, map_location="cpu")
            if "backbone" in sd:
                self.trainer.backbone.load_state_dict(sd["backbone"], strict=False)
            if "head" in sd:
                self.trainer.head.load_state_dict(sd["head"], strict=False)
        self.trainer.backbone.eval()
        self.trainer.head.eval()

    @torch.no_grad()
    def infer_study(self, dicom_dir: str) -> Dict[str, Any]:
        gate = self.chest_gate.is_chest_ct(dicom_dir)
        

        if not gate["is_chest"]:
            return {
                "prob": 0.0,
                "attn_weights": [],
                "cam_3d": None,
                "slices_gray": None,
                "meta": {}
            }

        # 1) полный проход: признаки, карты, веса
        logits, w, fmaps, slices_for_vis, meta = self.trainer._forward_patient_with_maps(dicom_dir)
        prob = float(torch.sigmoid(logits)[0, 0].item())

        # 2) CAM по весам классификатора
        W_cls = self.trainer.head.cls[1].weight.detach().cpu()  # [1, D] для бинарного
        w_pos = W_cls[0]                                        # [D]
        cam = torch.einsum('d,zdhw->zhw', w_pos, fmaps.cpu())   # [Z,h,w]
        cam = torch.relu(cam)
        cam = (cam - cam.amin(dim=(1,2), keepdim=True)) / (cam.amax(dim=(1,2), keepdim=True)
               - cam.amin(dim=(1,2), keepdim=True) + 1e-6)
        cam_3d = cam.numpy().astype(np.float32)                 # [Z,h,w] 0..1

        # 3) серые срезы для оверлея
        Z = len(slices_for_vis)
        slices_gray = None
        if Z > 0:
            H, W = slices_for_vis[0].shape
            slices_gray = np.stack([slices_for_vis[z] for z in range(Z)], axis=0).astype(np.float32)  # [Z,H,W] 0..1

        attn_weights = w.detach().cpu().numpy().tolist()
        
        return {
            "prob": prob,
            "attn_weights": attn_weights,
            "cam_3d": cam_3d,
            "slices_gray": slices_gray,
            "meta": meta
        }

class ModelZoo:
    def __init__(self, models: List[Tuple[str, str]], device="cuda", profile="balanced"):
        self.adapters = {name: ModelAdapter(device=device, ckpt_path=ckpt, profile=profile)
                         for name, ckpt in models}

    def infer_one_patient_row(self, dicom_dir: str, thr: float = 0.5) -> Dict[str, Any]:
        """
        Возвращает агрегированную строку для Excel по одному пациенту.
        """
        study_uid, series_uid = read_uid_pair(dicom_dir)

        start = time.time()
        per_model: Dict[str, Dict[str, Any]] = {}
        status = "Success"
        last_error = None

        try:
            for name, adapter in self.adapters.items():
                res = adapter.infer_study(dicom_dir)
                prob = float(res.get("prob", 0.0))
                attn = res.get("attn_weights", None)
                cam3d = res.get("cam_3d", None)
                slices_gray = res.get("slices_gray", None)
                meta = res.get('meta', {})

                # fallback для attn, если не пришёл
                if attn is None:
                    try:
                        _, w, _, _, meta = adapter.trainer._forward_patient_with_maps(dicom_dir)
                        attn = w.detach().cpu().numpy().tolist()
                    except Exception:
                        attn = None
                        meta = {}

                mask_soft = build_soft_mask(cam3d, slices_gray, attn)


                bbox, bbox_mm = robust_bbox_from_soft_mask(
                    mask_soft, 
                    spacing_xyz=meta.get("spacing", None),
                    thr=None, 
                    mode="percentile", 
                    pctl=99.0,
                    k_std=0.5,
                    smooth_sigma=1.0,
                    min_voxels=200,
                    ensure_nonempty=True
                )

                spacing_mm = None
                try:
                    # берём первую серию
                    series = collect_series(dicom_dir)
                    any_uid = next(iter(series))
                    some_ds = series[any_uid][0][1]
                    # sitk spacing — (x,y), z по шагу между срезами; в DICOM может отличаться
                    sx = float(getattr(some_ds, "PixelSpacing", [1.0,1.0])[0])
                    sy = float(getattr(some_ds, "PixelSpacing", [1.0,1.0])[1])
                    sz = float(getattr(some_ds, "SliceThickness", 1.0))
                    spacing_mm = (sx, sy, sz)
                except Exception as e:
                    print(e)


                meta_mask = {
                    "study_uid": study_uid,
                    "series_uid": series_uid,
                    "spacing_mm": spacing_mm,
                }
                mask_b64 = _pack_mask_b64(mask_soft, meta_mask)

                per_model[name] = {
                    "prob": prob,
                    "pred": int(prob >= thr),
                    "mask_b64": mask_b64,
                    "bbox": bbox,              # воксельные координаты
                    "bbox_mm": bbox_mm,
                }

        except Exception as e:
            status = "Failure"
            last_error = str(e)

        tproc = time.time() - start

        # выбор «самой опасной» патологии = модель с максимальной вероятностью
        most_name = None
        most_prob = -1.0

        for name, info in per_model.items():
            if info["prob"] > most_prob:
                most_name, most_prob = name, info["prob"]
        most_bbox = per_model[most_name]["bbox"] if most_name else None

        row: Dict[str, Any] = {
            "path_to_study": dicom_dir.replace("\\", "/"),
            "study_uid": str(study_uid or "unknown"),
            "series_uid": str(series_uid or "unknown"),
            "probability_of_pathology": float(max(0.0, most_prob)) if most_name else 0.0,
            "pathology": int(1 if (most_name and most_prob >= thr) else 0),
            "processing_status": "Success" if status == "Success" else "Failure",
            "time_of_processing": float(tproc),
            "most_dangerous_pathology_type": str(most_name or ""),
            "pathology_localization": json.dumps(most_bbox or []),  # как Float Array (строкой)
        }

        # полезные дополнительные колонки (по всем моделям)
        for name, info in per_model.items():
            row[f"prob@{name}"] = float(info["prob"])
            row[f"pred@{name}"] = int(info["pred"])
            row[f"mask_b64@{name}"] = info["mask_b64"]
            if info["bbox"] is not None:
                row[f"bbox_vox@{name}"] = json.dumps([float(x) for x in info["bbox"]])
            if info["bbox_mm"] is not None:
                row[f"bbox_mm@{name}"] = json.dumps([float(x) for x in info["bbox_mm"]])

        if status != "Success":
            row["error"] = last_error or ""
        return row


def make_default_cfg() -> FullConfig:

    cfg = FullConfig()

    cfg.model.backbone_type = "vit_2d"   # "vit_2d" | "resnet3d_medicalnet" | "cnn3d_custom"
    cfg.inp.use_25d = True
    cfg.inp.k_25d = 5
    cfg.prep.windows = [(-600,1500), (40,400)]
    cfg.model.in_chans = len(cfg.prep.windows) * cfg.inp.k_25d
    cfg.model.num_classes = 1

    return cfg

def _read_series_headers(dicom_dir):
    """Читает только заголовки, группирует по SeriesInstanceUID."""
    series = {}
    for root, _, files in os.walk(dicom_dir):
        for fn in files:
            path = os.path.join(root, fn)
            try:
                ds = pydicom.dcmread(path, stop_before_pixels=True, force=True)
            except Exception:
                continue
            uid = getattr(ds, "SeriesInstanceUID", None)
            if uid:
                series.setdefault(uid, []).append((path, ds))
    return series

def _best_series_paths(series_dict):
    """Выбираем «лучшую» серию: CT, монохром, больше срезов."""
    def is_ct(ds):
        return getattr(ds, "Modality", "") == "CT"
    def is_mono(ds):
        return int(getattr(ds, "SamplesPerPixel", 1)) == 1 and \
               getattr(ds, "PhotometricInterpretation", "MONOCHROME2").startswith("MONOCHROME")
    scored = []
    for uid, lst in series_dict.items():
        good = [(p,ds) for p,ds in lst if is_ct(ds) and is_mono(ds)]
        if len(good) >= 5:
            scored.append((uid, len(good), good))
    if not scored:
        # fallback: серия с макс. числом файлов
        uid = max(series_dict, key=lambda k: len(series_dict[k]))
        return [p for p,_ in series_dict[uid]]
    uid = max(scored, key=lambda t: t[1])[0]
    return [p for p,_ in series_dict[uid]]

def _read_image(paths):
    """Собираем серию SimpleITK (автоматически применяет slope/intercept → HU)."""
    reader = sitk.ImageSeriesReader()
    reader.SetFileNames(paths)
    return reader.Execute()

def _sitk_to_np(img):
    arr = sitk.GetArrayFromImage(img).astype(np.float32)  # [Z,Y,X] в HU
    sp  = img.GetSpacing()    # (sx, sy, sz)
    sz  = img.GetSize()       # (X, Y, Z)
    return arr, sp, sz

def _body_mask(vol):
    mask = vol > -500
    mask = ndi.binary_closing(mask, iterations=2)
    mask = ndi.binary_fill_holes(mask)
    lab, n = ndi.label(mask)
    if n > 0:
        sizes = ndi.sum(mask, lab, index=np.arange(1,n+1))
        k = 1 + int(np.argmax(sizes))
        mask = (lab == k)
    return mask

def _lung_mask(vol, body):
    lungs = (vol < -350) & body
    lungs = ndi.binary_opening(lungs, iterations=2)
    lungs = ndi.binary_closing(lungs, iterations=1)
    return lungs

def _two_lungs_fraction(lungs_slice):
    """Оцениваем, разделяется ли маска на 2 крупных компонента (левое/правое лёгкое)."""
    lab, n = ndi.label(lungs_slice)
    if n < 2:
        return 0.0
    sizes = ndi.sum(lungs_slice, lab, index=np.arange(1, n+1))
    sizes = np.sort(sizes)[::-1]
    if len(sizes) < 2:
        return 0.0
    s1, s2 = sizes[0], sizes[1]
    total = lungs_slice.sum() + 1e-6
    return float((s1 + s2) / total)

def safe_read_header(path: str):
    try:
        return pydicom.dcmread(path, stop_before_pixels=True, force=True)
    except Exception:
        return None

def collect_series(dicom_dir: str) -> dict:
    """Сканируем р+екурсивно все файлы, группируем по SeriesInstanceUID."""
    series = {}
    for root, _, fnames in os.walk(dicom_dir):
        for n in fnames:
            p = os.path.join(root, n)
            ds = safe_read_header(p)
            if ds is None:
                continue
            uid = getattr(ds, "SeriesInstanceUID", None)
            if uid is None:
                continue
            series.setdefault(uid, []).append((p, ds))
    return series

def build_soft_mask(cam_3d: np.ndarray,
                     slices_gray: Optional[np.ndarray],
                     attn_weights: Optional[List[float]]) -> np.ndarray:
    """
    cam_3d: [Z,h,w] в 0..1; slices_gray: [Z,H,W] (для целевого размера); attn_weights: [Z] или None
    return: [Z,H,W] float32 в 0..1
    """
    if cam_3d is None or not isinstance(cam_3d, np.ndarray) or cam_3d.ndim != 3:
        return None
    Z, h, w = cam_3d.shape
    if slices_gray is None:
        # возьмём H,W = h,w
        H, W = h, w
    else:
        H, W = slices_gray.shape[1], slices_gray.shape[2] if slices_gray.ndim==3 else slices_gray.shape[-1]

    cam = cam_3d.astype(np.float32).copy()
    # нормировка по срезу
    cam = (cam - cam.min(axis=(1,2), keepdims=True)) / (cam.max(axis=(1,2), keepdims=True) - cam.min(axis=(1,2), keepdims=True) + 1e-6)

    if attn_weights is not None and len(attn_weights) == Z:
        wv = np.asarray(attn_weights, np.float32)
        wv = wv / (wv.sum() + 1e-6)
        cam = cam * wv[:, None, None]

    # апскейл до [H,W]
    mask = np.zeros((Z, H, W), np.float32)
    for z in range(Z):
        m = cv2.resize(cam[z], (W, H), interpolation=cv2.INTER_CUBIC)
        mask[z] = np.clip(m, 0.0, 1.0)
    return mask

def _pack_mask_b64(mask_soft: Optional[np.ndarray], meta: Dict[str, Any]) -> str:
    if mask_soft is None: return ""
    arr = (np.clip(mask_soft, 0, 1) * 255.0).astype(np.uint8)
    bio = io.BytesIO()
    np.savez_compressed(bio, mask=arr, meta=np.array([meta], dtype=object))
    raw = bio.getvalue()
    return base64.b64encode(zlib.compress(raw, 9)).decode("ascii")

def robust_bbox_from_soft_mask(
    mask_soft: np.ndarray,
    spacing_xyz: tuple | None = None, 
    thr: float | None = None,
    mode: str = "percentile",
    pctl: float = 99.0,
    k_std: float = 0.5,
    smooth_sigma: float = 1.0,
    min_voxels: int = 200,
    ensure_nonempty: bool = True,
):
    m = mask_soft.astype(np.float32)

    # 1) нормализация к [0,1] по всему объёму
    vmin, vmax = float(m.min()), float(m.max())
    if vmax > vmin:
        m = (m - vmin) / (vmax - vmin)
    else:
        # вся маска константа → нечего локализовать
        if not ensure_nonempty:
            return None, None
        zc, yc, xc = 0, 0, 0
        return [0,0,0,0,0,0], [0,0,0,0,0,0] if spacing_xyz else None

    # 2) лёгкое сглаживание (устраняет шум)
    if smooth_sigma and smooth_sigma > 0:
        m = gaussian_filter(m, sigma=smooth_sigma)

    # 3) выбор порога
    if thr is None:
        if mode == "percentile":
            thr = float(np.percentile(m, pctl))
        elif mode == "mean+std":
            thr = float(m.mean() + k_std * m.std())
        elif mode == "otsu":
            # простой otsu по всему объёму
            hist, bins = np.histogram(m, bins=256, range=(0.0, 1.0))
            # otsu:
            w0 = 0; sum0 = 0; sum_total = (hist * ((bins[:-1] + bins[1:]) * 0.5)).sum()
            weight_total = hist.sum()
            max_between = -1; thr_bin = 0
            for i, h in enumerate(hist):
                w0 += h
                if w0 == 0 or w0 == weight_total: 
                    continue
                sum0 += h * ((bins[i] + bins[i+1]) * 0.5)
                m0 = sum0 / w0
                m1 = (sum_total - sum0) / (weight_total - w0)
                between = w0 * (weight_total - w0) * (m0 - m1) ** 2
                if between > max_between:
                    max_between = between
                    thr_bin = i
            thr = (bins[thr_bin] + bins[thr_bin+1]) * 0.5
        

    # 4) бинаризация
    bw = (m >= thr)

    # 5) удалим мелкие компоненты, возьмём крупнейшую
    lbl, n = label(bw)
    if n > 0:
        sizes = np.bincount(lbl.ravel())[1:]  # без фона
        idx_sorted = np.argsort(sizes)[::-1]
        kept = np.zeros_like(bw, dtype=bool)
        for i in idx_sorted:
            if sizes[i] >= min_voxels:
                kept |= (lbl == (i + 1))
                break
        if not kept.any():
            kept = (lbl == (idx_sorted[0] + 1))  # хотя бы крупнейшую
    else:
        kept = bw

    if not kept.any():
        if not ensure_nonempty:
            return None, None
        # 6) «минимальный бокс»: вокруг глобального максимума, радиус 1 воксель
        zc, yc, xc = np.unravel_index(np.argmax(m), m.shape)
        z0, z1 = max(0, zc-1), min(m.shape[0]-1, zc+1)
        y0, y1 = max(0, yc-2), min(m.shape[1]-1, yc+2)
        x0, x1 = max(0, xc-2), min(m.shape[2]-1, xc+2)
        bbox_vox = [float(x0), float(x1), float(y0), float(y1), float(z0), float(z1)]
    else:
        zz, yy, xx = np.where(kept)
        z0, z1 = int(zz.min()), int(zz.max())
        y0, y1 = int(yy.min()), int(yy.max())
        x0, x1 = int(xx.min()), int(xx.max())
        bbox_vox = [float(x0), float(x1), float(y0), float(y1), float(z0), float(z1)]

    # 7) опционально — бокс в мм
    bbox_mm = None
    if spacing_xyz is not None:
        sx, sy, sz = spacing_xyz
        x0, x1, y0, y1, z0, z1 = bbox_vox
        bbox_mm = [x0*sx, x1*sx, y0*sy, y1*sy, z0*sz, z1*sz]

    return bbox_vox, bbox_mm

def _autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 12
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len: max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(80, max_len + 2)

def rows_to_xlsx_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """
    Собирает .xlsx в памяти и возвращает bytes.
    """

    required = [
        "path_to_study","study_uid","series_uid","probability_of_pathology",
        "pathology","processing_status","time_of_processing",
        "most_dangerous_pathology_type","pathology_localization"
    ]
    extra_keys = sorted({k for r in rows for k in r.keys() if k not in required})
    header = required + extra_keys

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(header)
    for r in rows:
        ws.append([r.get(k, "") for k in header])

    _autosize(ws)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def is_ct_monochrome(ds) -> bool:
    try:
        if getattr(ds, "Modality", "") != "CT":
            return False
        if int(getattr(ds, "SamplesPerPixel", 1)) != 1:
            return False
        if getattr(ds, "PhotometricInterpretation", "MONOCHROME2") not in ("MONOCHROME1","MONOCHROME2"):
            return False
        return True
    except Exception:
        return False

def series_score(files_with_hdrs: List[Tuple[str, Any]]) -> float:
    """
    Оцениваем серию: больше срезов, присутствуют slope/intercept, достаточная битность.
    Чем выше — тем лучше.
    """
    n = len(files_with_hdrs)
    # признаки
    slopes = 0
    bits = 0
    mono = 0
    for _, ds in files_with_hdrs:
        if hasattr(ds, "RescaleSlope") and hasattr(ds, "RescaleIntercept"):
            slopes += 1
        bits += int(getattr(ds, "BitsStored", 12) or 12)
        if int(getattr(ds, "SamplesPerPixel", 1)) == 1:
            mono += 1
    # простая взвешенная сумма
    return (n * 1.0) + (slopes / max(1, n)) * 0.5 + (mono / max(1, n)) * 0.2 + (bits / max(1, n)) * 0.01

def sort_by_z(files_with_hdrs: List[Tuple[str, Any]]) -> List[str]:
    def zpos(ds):
        ipp = getattr(ds, "ImagePositionPatient", None)
        if ipp is not None and len(ipp) == 3:
            return float(ipp[2])
        inst = getattr(ds, "InstanceNumber", None)
        return float(inst) if inst is not None else 0.0
    files = [(p, ds) for p,ds in files_with_hdrs if ds is not None]
    files.sort(key=lambda t: zpos(t[1]))
    return [p for p,_ in files]

def pick_best_ct_series(dicom_dir: str) -> Tuple[Optional[str], Optional[str], Optional[List[str]]]:
    """
    Returns (study_uid, series_uid, sorted_file_list) or (None,None,None) if not found.
    """
    series = collect_series(dicom_dir)
    if not series:
        return None, None, None

    candidates = {}
    for uid, files in series.items():
        filt = [(p, ds) for (p, ds) in files if is_ct_monochrome(ds)]
        if len(filt) >= 3:
            candidates[uid] = filt

    if not candidates:
        # fallback: largest series, but warn by returning still
        best_uid = max(series.keys(), key=lambda u: len(series[u]))
        files = series[best_uid]
    else:
        best_uid = max(candidates.keys(), key=lambda u: series_score(candidates[u]))
        files = candidates[best_uid]

    sorted_files = sort_by_z(files)
    if not sorted_files:
        return None, None, None

    ds0 = files[0][1]
    study_uid = getattr(ds0, "StudyInstanceUID", "unknown")
    return study_uid, best_uid, sorted_files

def read_uid_pair(dicom_dir: str) -> Tuple[str,str]:
    s, u, _ = pick_best_ct_series(dicom_dir)
    return s or "unknown_study", u or "unknown_series"

def _is_dicom_file(path: str) -> bool:
    try:
        if not os.path.isfile(path):
            return False
        with open(path, 'rb') as f:
            pre = f.read(132)
            if len(pre) >= 132 and pre[128:132] == b'DICM':
                return True

        ds = safe_read_header(path)
        return ds is not None
    except Exception:
        return False

def _dir_has_dicoms(dir_path: str) -> bool:
    for root, _, files in os.walk(dir_path):
        for n in files:
            p = os.path.join(root, n)
            if _is_dicom_file(p):
                return True
    return False

def _list_patient_dirs(root: str) -> List[str]:
    """
    Правила:
    - если root сам содержит DICOM → это один пациент;
    - иначе ищем подкаталоги, которые сами содержат DICOM → каждый = пациент;
    - если среди подкаталогов нет DICOM, но есть ещё уровень — углубляемся (1 уровень).
    """
    root = os.path.abspath(root)
    if _dir_has_dicoms(root):
        return [root]

    candidates = []
    subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]
    for d in subdirs:
        if _dir_has_dicoms(d):
            candidates.append(d)

    if candidates:
        return sorted(candidates)

    for d in subdirs:
        for d2 in [os.path.join(d, x) for x in os.listdir(d) if os.path.isdir(os.path.join(d, x))]:
            if _dir_has_dicoms(d2):
                candidates.append(d2)

    return sorted(candidates)

def main():
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    profile = 'fast_debug'
    thr = 0.55
    input_path = r"" # путь к zip-архиву

    temp_dir = None
    work_root = input_path
    if os.path.isfile(input_path) and input_path.lower().endswith(".zip"):
        temp_dir = tempfile.mkdtemp(prefix="ctzip_")
        with zipfile.ZipFile(input_path, "r") as zf:
            zf.extractall(temp_dir)
        work_root = temp_dir

    patient_dirs = _list_patient_dirs(work_root)
    if not patient_dirs:
        subdirs = [os.path.join(work_root, d) for d in os.listdir(work_root) if os.path.isdir(os.path.join(work_root, d))]
        for d in subdirs:
            patient_dirs.extend(_list_patient_dirs(d))
        patient_dirs = sorted(set(patient_dirs))

    if not patient_dirs:
        rows: List[Dict[str, Any]] = []
        xlsx_bytes = rows_to_xlsx_bytes(rows)
        if temp_dir: shutil.rmtree(temp_dir, ignore_errors=True)
        return xlsx_bytes

    zoo = ModelZoo(MODELS, device=device, profile=profile)
    rows = []
    for pdir in patient_dirs:
        row = zoo.infer_one_patient_row(pdir, thr=thr)
        rows.append(row)

    xlsx_bytes = rows_to_xlsx_bytes(rows)
    if temp_dir: shutil.rmtree(temp_dir, ignore_errors=True)

    return xlsx_bytes

if __name__ == "__main__":
    main()
